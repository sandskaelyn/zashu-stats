from flask import Flask, request, send_file, render_template, jsonify
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta
import os
import tempfile
import json

app = Flask(__name__)

THIN_BORDER = Border(
    top=Side(style='thin'),
    bottom=Side(style='thin'),
    left=Side(style='thin'),
    right=Side(style='thin')
)

BOTTOM_LR_BORDER = Border(
    top=Side(style=None),
    bottom=Side(style='thin'),
    left=Side(style='thin'),
    right=Side(style='thin')
)

TITLE_FONT = Font(name='微软雅黑 Light', size=20, bold=True)
HEADER_FONT = Font(name='微软雅黑 Light', size=18, bold=False)
DATA_FONT = Font(name='微软雅黑 Light', size=18, bold=False)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')

COLUMN_WIDTHS = {
    'A': 9.0, 'C': 13.5, 'D': 9.375, 'E': 9.375, 'F': 9.375,
    'G': 10.125, 'H': 13.0, 'I': 9.375, 'J': 11.25
}


def parse_excel_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        epoch = datetime(1899, 12, 30)
        dt = epoch + timedelta(days=int(v))
        return dt.date()
    return None


def generate_excel(detail_path, year, month, locations, filter_baogan):
    wb_detail = openpyxl.load_workbook(detail_path, data_only=True)
    ws_detail = wb_detail.active

    header = [cell.value for cell in next(ws_detail.iter_rows(min_row=1, max_row=1))]
    col_date = next((i for i, h in enumerate(header) if h and '日期' in str(h)), -1)
    col_zashu = next((i for i, h in enumerate(header) if h and '销售扎数' in str(h)), -1)
    col_baogan = next((i for i, h in enumerate(header) if h and '包干费代收' in str(h)), -1)
    col_addr = next((i for i, h in enumerate(header) if h and '地址' in str(h)), -1)
    col_site = next((i for i, h in enumerate(header) if h and '二级站点' in str(h)), -1)

    if col_date == -1 or col_zashu == -1:
        raise ValueError('未找到日期或销售扎数列')

    day_map = {}
    total_rows = 0
    baogan_filtered = 0
    location_matched = 0

    for row in ws_detail.iter_rows(min_row=2, values_only=True):
        raw_date = row[col_date]
        d = parse_excel_date(raw_date)
        if not d:
            continue
        total_rows += 1

        if d.year != year or d.month != month:
            continue

        if filter_baogan:
            baogan_val = row[col_baogan] if col_baogan >= 0 else None
            if baogan_val == 0 or baogan_val == '0' or baogan_val is None:
                baogan_filtered += 1
                continue

        zashu_val = row[col_zashu] or 0
        if not zashu_val:
            continue

        address_val = str(row[col_addr] or '') if col_addr >= 0 else ''
        site_val = str(row[col_site] or '') if col_site >= 0 else ''
        combined = address_val + site_val

        day = d.day
        if day not in day_map:
            day_map[day] = [0] * len(locations)

        for li, loc in enumerate(locations):
            if loc['match'] and loc['match'] in combined:
                day_map[day][li] += zashu_val
                location_matched += 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '三地址汇总'

    loc_count = len(locations)
    days_in_month = (datetime(year, month + 1, 1) - timedelta(days=1)).day

    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 15
    ws.row_dimensions[2].height = 45
    ws.merge_cells('C2:J2')
    title_cell = ws['C2']
    title_cell.value = f'黄冈仓{month}月份二级站点扎数总计'
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER_ALIGN
    title_cell.border = THIN_BORDER
    for c in range(4, 11):
        ws.cell(row=2, column=c).border = THIN_BORDER

    ws.row_dimensions[3].height = 50
    headers = ['日期'] + [loc['name'] for loc in locations] + ['合计']
    header_cols = [3, 4, 5, 6, 10]
    for i, col_idx in enumerate(header_cols):
        if i < len(headers):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = headers[i]
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

    data_start_row = 4
    data_end_row = data_start_row + days_in_month - 1
    col_totals = [0] * loc_count
    grand_total = 0

    for day in range(1, days_in_month + 1):
        row_idx = data_start_row + day - 1
        ws.row_dimensions[row_idx].height = 24.75
        vals = day_map.get(day, [0] * loc_count)
        day_total = sum(vals)
        grand_total += day_total

        date_cell = ws.cell(row=row_idx, column=3)
        date_cell.value = datetime(year, month, day)
        date_cell.number_format = 'm/d'
        date_cell.font = DATA_FONT
        date_cell.alignment = CENTER_ALIGN
        date_cell.border = BOTTOM_LR_BORDER

        for li in range(loc_count):
            cell = ws.cell(row=row_idx, column=4 + li)
            cell.value = vals[li] if vals[li] > 0 else None
            cell.font = DATA_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = BOTTOM_LR_BORDER
            col_totals[li] += vals[li]

        total_cell = ws.cell(row=row_idx, column=10)
        total_cell.value = day_total if day_total > 0 else None
        total_cell.font = DATA_FONT
        total_cell.alignment = CENTER_ALIGN
        total_cell.border = BOTTOM_LR_BORDER

    total_row = data_end_row + 1
    ws.row_dimensions[total_row].height = 24.75

    total_label = ws.cell(row=total_row, column=3)
    total_label.value = '合计'
    total_label.font = DATA_FONT
    total_label.alignment = CENTER_ALIGN
    total_label.border = BOTTOM_LR_BORDER

    for li in range(loc_count):
        cell = ws.cell(row=total_row, column=4 + li)
        cell.value = col_totals[li] if col_totals[li] > 0 else None
        cell.font = DATA_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = BOTTOM_LR_BORDER

    total_cell = ws.cell(row=total_row, column=10)
    total_cell.value = grand_total if grand_total > 0 else None
    total_cell.font = DATA_FONT
    total_cell.alignment = CENTER_ALIGN
    total_cell.border = BOTTOM_LR_BORDER

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()

    return tmp.name, {
        'total_rows': total_rows,
        'baogan_filtered': baogan_filtered,
        'location_matched': location_matched,
        'grand_total': grand_total
    }


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/generate', methods=['POST'])
def api_generate():
    try:
        file = request.files.get('file')
        if not file:
            return jsonify({'error': '请上传明细表'}), 400

        year = int(request.form.get('year', 0))
        month = int(request.form.get('month', 0))
        locations = json.loads(request.form.get('locations', '[]'))
        filter_baogan = request.form.get('filter_baogan', 'false') == 'true'

        if not year or not month:
            return jsonify({'error': '请选择月份'}), 400
        if not locations:
            return jsonify({'error': '请至少配置一个地点'}), 400

        tmp_detail = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        file.save(tmp_detail.name)
        tmp_detail.close()

        output_path, stats = generate_excel(
            tmp_detail.name, year, month, locations, filter_baogan
        )

        os.unlink(tmp_detail.name)

        filename = f'{year}年黄冈仓{month}月二级站点扎数统计.xlsx'
        return send_file(output_path, as_attachment=True, download_name=filename)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(host='0.0.0.0', port=port, debug=debug)
